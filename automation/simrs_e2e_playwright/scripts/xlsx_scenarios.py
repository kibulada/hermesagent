#!/usr/bin/env python3
"""
xlsx_scenarios.py — baca sheet skenario test (.xlsx) jadi JSON terstruktur.

Sumber AC di tim Kesia tidak selalu tiket OpenProject. Sebagian datang sebagai
sheet Excel (mis. `SATUSEHAT_PARSIAL_TEST_SCENARIO_filled.xlsx`) yang sebelumnya
tidak pernah dibaca kode mana pun — hanya dibuka manual.

Bentuk sheet yang ditangani (kolom Sheet1):
    No | Task | Module | Scenario | Expectation | Evidence | Progress | Dev Status | Status | Notes

Dua hal yang bikin parsing naif salah:
  1. **Merged cell** — `No` dan `Task` dikosongkan saat baris melanjutkan grup di
     atasnya. Tanpa forward-fill, modul kehilangan Task-nya.
  2. **Beberapa case dalam satu sel** — `Scenario` dan `Expectation` memuat blok
     "Positive Case:" dan "Negative Case:" sekaligus, dipisah baris baru.
     Diperlakukan sebagai satu skenario akan menggabungkan dua kasus yang berlawanan.

Usage:
    python scripts/xlsx_scenarios.py --file ../../SATUSEHAT_PARSIAL_TEST_SCENARIO_filled.xlsx
    python scripts/xlsx_scenarios.py --file <path> --sheet Sheet1 --module-filter Registrasi
"""
import argparse
import json
import re
import sys
from pathlib import Path
from typing import Optional

try:
    import openpyxl
except ImportError:
    raise SystemExit('openpyxl belum terpasang. pip install -r requirements.txt')

# kolom yang di-forward-fill saat sel kosong (efek merged cell)
CARRY_COLUMNS = ('no', 'task')

CASE_SPLIT = re.compile(r'^\s*(Positive|Negative)\s+Case\s*:\s*$', re.I | re.M)


def norm_header(value) -> str:
    return re.sub(r'[^a-z0-9]+', '_', str(value or '').strip().lower()).strip('_')


def split_cases(text: str) -> dict:
    """
    Pecah 'Positive Case:\\n1. ...\\nNegative Case:\\n1. ...' jadi {'positive': ..., 'negative': ...}.
    Teks tanpa penanda case dikembalikan sebagai {'general': teks}.
    """
    if not text or not text.strip():
        return {}

    parts = CASE_SPLIT.split(text)
    if len(parts) == 1:
        return {'general': text.strip()}

    result = {}
    # parts = [prefix, 'Positive', body, 'Negative', body, ...]
    prefix = parts[0].strip()
    if prefix:
        result['general'] = prefix
    for label, body in zip(parts[1::2], parts[2::2]):
        body = body.strip()
        if body:
            result[label.lower()] = body
    return result


def steps_of(text: str) -> list:
    """Ambil langkah bernomor ('1. ...') kalau ada; kalau tidak, pecah per baris non-kosong."""
    if not text:
        return []
    numbered = re.findall(r'^\s*\d+[.)]\s*(.+)$', text, re.M)
    if numbered:
        return [s.strip() for s in numbered if s.strip()]
    return [ln.strip() for ln in text.splitlines() if ln.strip()]


def read_sheet(path: Path, sheet: Optional[str] = None) -> list:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    header = None
    for raw in rows:
        if raw and any(c is not None and str(c).strip() for c in raw):
            header = [norm_header(c) for c in raw]
            break
    if not header:
        return []

    out = []
    carry = {}
    for raw in rows:
        if not raw or not any(c is not None and str(c).strip() for c in raw):
            continue
        row = {}
        for key, value in zip(header, raw):
            if not key:
                continue
            row[key] = str(value).strip() if value is not None else ''

        # forward-fill efek merged cell
        for key in CARRY_COLUMNS:
            if row.get(key):
                carry[key] = row[key]
            elif carry.get(key):
                row[key] = carry[key]

        if not row.get('module') and not row.get('scenario'):
            continue

        scenarios = split_cases(row.get('scenario', ''))
        expectations = split_cases(row.get('expectation', ''))

        cases = []
        for kind in sorted(set(scenarios) | set(expectations)):
            cases.append({
                'kind': kind,                                   # positive | negative | general
                'steps': steps_of(scenarios.get(kind, '')),
                'expected': expectations.get(kind, ''),
            })

        out.append({
            'no': row.get('no', ''),
            'task': row.get('task', ''),
            'module': row.get('module', ''),
            'cases': cases,
            'status': row.get('status', ''),
            'dev_status': row.get('dev_status', ''),
            'notes': row.get('notes', ''),
        })

    wb.close()
    return out


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument('--file', required=True, help='path file .xlsx')
    p.add_argument('--sheet', default=None, help='nama sheet (default: sheet pertama)')
    p.add_argument('--module-filter', default='', help='hanya modul yang mengandung teks ini')
    p.add_argument('--list-sheets', action='store_true', help='tampilkan nama sheet lalu keluar')
    args = p.parse_args()

    path = Path(args.file)
    if not path.exists():
        raise SystemExit(f'File tidak ditemukan: {path}')

    if args.list_sheets:
        wb = openpyxl.load_workbook(path, read_only=True)
        print(json.dumps({'sheets': wb.sheetnames}, indent=2))
        wb.close()
        return 0

    rows = read_sheet(path, args.sheet)
    if args.module_filter:
        needle = args.module_filter.lower()
        rows = [r for r in rows if needle in r['module'].lower()]

    print(json.dumps({
        'source': str(path),
        'sheet': args.sheet or 'default',
        'total': len(rows),
        'scenarios': rows,
    }, indent=2, ensure_ascii=False))
    return 0


if __name__ == '__main__':
    sys.exit(main())
